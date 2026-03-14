from flask import Flask, request, jsonify, send_from_directory
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os, hashlib, datetime, re

app = Flask(__name__, static_folder='static', template_folder='templates')
DB = 'students_db.xlsx'

COLS_USERS = ['ID','Name','Email','Password','Role','CreatedAt']
COLS_STUDENTS = ['StudentID','Name','Email','Phone','Course','Year','GPA','Status','EnrolledDate','Address','Guardian']
COLS_ATTENDANCE = ['RecordID','StudentID','Date','Status','Subject']
COLS_GRADES = ['GradeID','StudentID','Subject','Marks','MaxMarks','Grade','Semester']
COLS_FEES = ['FeeID','StudentID','Amount','PaidDate','DueDate','Status','Description']

def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def init_db():
    if os.path.exists(DB):
        return
    wb = Workbook()
    def make_sheet(wb, name, cols, first=False):
        ws = wb.active if first else wb.create_sheet(name)
        if first: ws.title = name
        hdr_fill = PatternFill('solid', start_color='1E3A5F')
        hdr_font = Font(bold=True, color='FFFFFF', size=11)
        for i, c in enumerate(cols, 1):
            cell = ws.cell(1, i, c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center')
        ws.freeze_panes = 'A2'
        return ws
    make_sheet(wb, 'Users', COLS_USERS, first=True)
    make_sheet(wb, 'Students', COLS_STUDENTS)
    make_sheet(wb, 'Attendance', COLS_ATTENDANCE)
    make_sheet(wb, 'Grades', COLS_GRADES)
    make_sheet(wb, 'Fees', COLS_FEES)
    # default admin
    wb['Users'].append(['USR001','Admin','admin@sms.com', hash_pw('admin123'),'admin', str(datetime.date.today())])
    wb.save(DB)

def get_sheet(name):
    wb = load_workbook(DB)
    return wb, wb[name]

def sheet_to_list(ws):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2: return []
    headers = rows[0]
    return [dict(zip(headers, r)) for r in rows[1:] if any(r)]

def next_id(ws, prefix):
    rows = list(ws.iter_rows(min_row=2, max_col=1, values_only=True))
    nums = [int(str(r[0]).replace(prefix,'')) for r in rows if r[0] and str(r[0]).startswith(prefix)]
    return f"{prefix}{(max(nums)+1 if nums else 1):03d}"

init_db()

# ── Serve HTML files ──────────────────────────────────────────
@app.route('/')
def index(): return send_from_directory('templates','login.html')

@app.route('/<path:fn>')
def static_files(fn):
    for d in ['templates','static']:
        p = os.path.join(d, fn)
        if os.path.exists(p): return send_from_directory(d, fn)
    return 'Not found', 404

# ── AUTH ──────────────────────────────────────────────────────
@app.route('/api/login', methods=['POST'])
def login():
    d = request.json
    wb, ws = get_sheet('Users')
    users = sheet_to_list(ws)
    user = next((u for u in users if u['Email']==d.get('email') and u['Password']==hash_pw(d.get('password',''))), None)
    if not user: return jsonify({'ok':False,'msg':'Invalid credentials'}), 401
    return jsonify({'ok':True,'user':{'id':user['ID'],'name':user['Name'],'email':user['Email'],'role':user['Role']}})

@app.route('/api/signup', methods=['POST'])
def signup():
    d = request.json
    wb, ws = get_sheet('Users')
    users = sheet_to_list(ws)
    if any(u['Email']==d.get('email') for u in users):
        return jsonify({'ok':False,'msg':'Email already registered'}), 400
    uid = next_id(ws,'USR')
    ws.append([uid, d['name'], d['email'], hash_pw(d['password']), d.get('role','student'), str(datetime.date.today())])
    wb.save(DB)
    return jsonify({'ok':True,'msg':'Account created!'})

# ── STUDENTS ──────────────────────────────────────────────────
@app.route('/api/students', methods=['GET'])
def get_students():
    _, ws = get_sheet('Students')
    return jsonify(sheet_to_list(ws))

@app.route('/api/students', methods=['POST'])
def add_student():
    d = request.json
    wb, ws = get_sheet('Students')
    sid = next_id(ws,'STU')
    ws.append([sid, d['name'], d['email'], d.get('phone',''), d.get('course',''), d.get('year',1),
               d.get('gpa',0.0), d.get('status','Active'), str(datetime.date.today()),
               d.get('address',''), d.get('guardian','')])
    wb.save(DB)
    return jsonify({'ok':True,'id':sid})

@app.route('/api/students/<sid>', methods=['PUT'])
def update_student(sid):
    d = request.json
    wb, ws = get_sheet('Students')
    for row in ws.iter_rows(min_row=2):
        if row[0].value == sid:
            fields = ['name','email','phone','course','year','gpa','status','enrolledDate','address','guardian']
            keys   = ['Name','Email','Phone','Course','Year','GPA','Status','EnrolledDate','Address','Guardian']
            for i,k in enumerate(keys):
                fk = fields[i]
                if fk in d: row[i+1].value = d[fk]
            wb.save(DB)
            return jsonify({'ok':True})
    return jsonify({'ok':False,'msg':'Not found'}), 404

@app.route('/api/students/<sid>', methods=['DELETE'])
def delete_student(sid):
    wb, ws = get_sheet('Students')
    for row in ws.iter_rows(min_row=2):
        if row[0].value == sid:
            ws.delete_rows(row[0].row)
            wb.save(DB)
            return jsonify({'ok':True})
    return jsonify({'ok':False,'msg':'Not found'}), 404

# ── ATTENDANCE ────────────────────────────────────────────────
@app.route('/api/attendance', methods=['GET'])
def get_attendance():
    _, ws = get_sheet('Attendance')
    return jsonify(sheet_to_list(ws))

@app.route('/api/attendance', methods=['POST'])
def add_attendance():
    d = request.json
    wb, ws = get_sheet('Attendance')
    rid = next_id(ws,'ATT')
    ws.append([rid, d['studentId'], d.get('date', str(datetime.date.today())), d['status'], d.get('subject','')])
    wb.save(DB)
    return jsonify({'ok':True})

# ── GRADES ────────────────────────────────────────────────────
@app.route('/api/grades', methods=['GET'])
def get_grades():
    _, ws = get_sheet('Grades')
    return jsonify(sheet_to_list(ws))

@app.route('/api/grades', methods=['POST'])
def add_grade():
    d = request.json
    wb, ws = get_sheet('Grades')
    gid = next_id(ws,'GRD')
    marks, max_m = float(d.get('marks',0)), float(d.get('maxMarks',100))
    pct = (marks/max_m)*100 if max_m else 0
    grade = 'A+' if pct>=95 else 'A' if pct>=85 else 'B+' if pct>=75 else 'B' if pct>=65 else 'C' if pct>=55 else 'D' if pct>=40 else 'F'
    ws.append([gid, d['studentId'], d['subject'], marks, max_m, grade, d.get('semester',1)])
    wb.save(DB)
    return jsonify({'ok':True})

# ── FEES ──────────────────────────────────────────────────────
@app.route('/api/fees', methods=['GET'])
def get_fees():
    _, ws = get_sheet('Fees')
    return jsonify(sheet_to_list(ws))

@app.route('/api/fees', methods=['POST'])
def add_fee():
    d = request.json
    wb, ws = get_sheet('Fees')
    fid = next_id(ws,'FEE')
    ws.append([fid, d['studentId'], float(d.get('amount',0)),
               d.get('paidDate',''), d.get('dueDate',''), d.get('status','Pending'), d.get('description','')])
    wb.save(DB)
    return jsonify({'ok':True})

# ── STATS ─────────────────────────────────────────────────────
@app.route('/api/stats')
def stats():
    _, ws_s = get_sheet('Students')
    _, ws_a = get_sheet('Attendance')
    _, ws_g = get_sheet('Grades')
    _, ws_f = get_sheet('Fees')
    students = sheet_to_list(ws_s)
    att = sheet_to_list(ws_a)
    grades = sheet_to_list(ws_g)
    fees = sheet_to_list(ws_f)
    active = sum(1 for s in students if s.get('Status')=='Active')
    present = sum(1 for a in att if a.get('Status')=='Present')
    att_rate = round((present/len(att))*100,1) if att else 0
    gpas = [float(s['GPA']) for s in students if s.get('GPA')]
    avg_gpa = round(sum(gpas)/len(gpas),2) if gpas else 0
    pending_fees = sum(float(f['Amount']) for f in fees if f.get('Status')=='Pending' and f.get('Amount'))
    return jsonify({'totalStudents':len(students),'activeStudents':active,
                    'attendanceRate':att_rate,'avgGPA':avg_gpa,'pendingFees':pending_fees})

if __name__=='__main__':
    app.run(debug=True, port=5000)
